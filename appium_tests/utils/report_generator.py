"""
utils/report_generator.py — Excel Report Generator
====================================================
Generates a rich, color-coded Excel report with:
  - Sheet 1: Executive Summary (pass/fail counts, duration, timestamp)
  - Sheet 2: Detailed Test Results (per-test row with status, error, screenshot)
  - Sheet 3: Screen Coverage Map
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config


# ── Color Palette ──────────────────────────────────────────────────────────────
COLOR_PASS        = "1DB954"   # Green
COLOR_FAIL        = "E53935"   # Red
COLOR_SKIP        = "F57C00"   # Orange
COLOR_HEADER_BG   = "0D1B2A"   # Dark navy (matches app theme)
COLOR_HEADER_FG   = "00E5FF"   # Neon Cyan (matches app theme)
COLOR_ROW_ALT     = "1A2535"   # Alternating row dark
COLOR_ROW_NORMAL  = "0F172A"   # Normal row dark
COLOR_TITLE_BG    = "001529"   # Deep dark
COLOR_BORDER      = "2A3F5F"   # Subtle border


def _thin_border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font():
    return Font(name="Calibri", bold=True, color=COLOR_HEADER_FG, size=11)


def _title_font(size=18):
    return Font(name="Calibri", bold=True, color=COLOR_HEADER_FG, size=size)


def _pass_fill():
    return PatternFill("solid", fgColor=COLOR_PASS)


def _fail_fill():
    return PatternFill("solid", fgColor=COLOR_FAIL)


def _skip_fill():
    return PatternFill("solid", fgColor=COLOR_SKIP)


def _header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)


def _alt_fill(i):
    color = COLOR_ROW_ALT if i % 2 == 0 else COLOR_ROW_NORMAL
    return PatternFill("solid", fgColor=color)


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def generate_excel_report(results: list, start_time: datetime, end_time: datetime) -> str:
    """
    Generate the Excel test report.

    Args:
        results: List of dicts with keys:
            - test_id (int)
            - test_name (str)
            - screen (str)
            - status (str): 'PASS' | 'FAIL' | 'SKIP'
            - duration (float): seconds
            - error (str): error message if failed
            - screenshot (str): path to screenshot file
        start_time: Test suite start datetime
        end_time: Test suite end datetime

    Returns:
        str: Path to the generated Excel file
    """
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "📊 Summary"
    _build_summary_sheet(ws_summary, results, start_time, end_time)

    # ── Sheet 2: Detailed Results ──────────────────────────────────────────────
    ws_detail = wb.create_sheet("🧪 Detailed Results")
    _build_detail_sheet(ws_detail, results)

    # ── Sheet 3: Screen Coverage ───────────────────────────────────────────────
    ws_coverage = wb.create_sheet("📱 Screen Coverage")
    _build_coverage_sheet(ws_coverage, results)

    # ── Save ───────────────────────────────────────────────────────────────────
    timestamp = end_time.strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"Nutrixa_E2E_Report_{timestamp}.xlsx"
    filepath = os.path.join(config.REPORTS_DIR, filename)
    wb.save(filepath)

    print(f"\n  📊 Excel report generated: {filepath}\n")
    return filepath


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 1 — SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, results, start_time, end_time):
    ws.sheet_view.showGridLines = False

    total = len(results)
    passed = sum(1 for r in results if r["status"] == "PASS")
    failed = sum(1 for r in results if r["status"] == "FAIL")
    skipped = sum(1 for r in results if r["status"] == "SKIP")
    duration = (end_time - start_time).total_seconds()
    pass_rate = (passed / total * 100) if total > 0 else 0

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "🧬 NUTRIXA (PONIS) — End-to-End Appium Test Report"
    title_cell.font = _title_font(16)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # ── Metadata ───────────────────────────────────────────────────────────────
    meta = [
        ("Run Date", start_time.strftime("%Y-%m-%d")),
        ("Start Time", start_time.strftime("%H:%M:%S")),
        ("End Time", end_time.strftime("%H:%M:%S")),
        ("Total Duration", f"{duration:.1f} seconds"),
        ("App Package", config.APP_PACKAGE),
        ("Device", config.DEVICE_NAME),
        ("Android Version", config.PLATFORM_VERSION),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws[f"A{i}"].value = label
        ws[f"A{i}"].font = Font(name="Calibri", bold=True, color="8899AA", size=10)
        ws[f"A{i}"].fill = PatternFill("solid", fgColor=COLOR_ROW_NORMAL)

        ws[f"B{i}"].value = value
        ws[f"B{i}"].font = Font(name="Calibri", color="FFFFFF", size=10)
        ws[f"B{i}"].fill = PatternFill("solid", fgColor=COLOR_ROW_NORMAL)

    # ── KPI Cards ──────────────────────────────────────────────────────────────
    kpi_start_row = 12

    kpi_data = [
        ("D", "TOTAL TESTS", total,  "2A3F5F", "FFFFFF"),
        ("E", "PASSED ✅",   passed,  COLOR_PASS,  "FFFFFF"),
        ("F", "FAILED ❌",   failed,  COLOR_FAIL,  "FFFFFF"),
        ("G", "SKIPPED ⏭️",  skipped, COLOR_SKIP,  "FFFFFF"),
        ("H", "PASS RATE",   f"{pass_rate:.1f}%", COLOR_HEADER_BG, COLOR_HEADER_FG),
    ]

    for col, label, value, bg, fg in kpi_data:
        label_cell = ws[f"{col}{kpi_start_row}"]
        value_cell = ws[f"{col}{kpi_start_row + 1}"]

        label_cell.value = label
        label_cell.font = Font(name="Calibri", bold=True, color=fg, size=9)
        label_cell.fill = PatternFill("solid", fgColor=bg)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        value_cell.value = value
        value_cell.font = Font(name="Calibri", bold=True, color=fg, size=22)
        value_cell.fill = PatternFill("solid", fgColor=bg)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[kpi_start_row].height = 22
        ws.row_dimensions[kpi_start_row + 1].height = 45

    # ── Pie Chart ──────────────────────────────────────────────────────────────
    chart_data_row = 20
    ws[f"D{chart_data_row}"].value = "Status"
    ws[f"E{chart_data_row}"].value = "Count"
    chart_rows = [("Passed", passed), ("Failed", failed), ("Skipped", skipped)]
    for i, (label, count) in enumerate(chart_rows, start=chart_data_row + 1):
        ws[f"D{i}"].value = label
        ws[f"E{i}"].value = count

    pie = PieChart()
    pie.title = "Test Result Distribution"
    pie.style = 10
    data_ref = Reference(ws, min_col=5, min_row=chart_data_row, max_row=chart_data_row + 3)
    labels_ref = Reference(ws, min_col=4, min_row=chart_data_row + 1, max_row=chart_data_row + 3)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    pie.dataLabels = None
    ws.add_chart(pie, "D15")

    _set_col_widths(ws, {
        "A": 20, "B": 28, "C": 5,
        "D": 18, "E": 18, "F": 18, "G": 18, "H": 18
    })


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 2 — DETAILED RESULTS
# ─────────────────────────────────────────────────────────────────────────────

def _build_detail_sheet(ws, results):
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "🧪 Detailed Test Results — All Screens"
    title_cell.font = _title_font(14)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers
    headers = ["#", "Test Name", "Screen", "Status", "Duration (s)", "Error Message", "Screenshot File", "Timestamp"]
    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[header_row].height = 28

    # Data rows
    for i, result in enumerate(results, start=1):
        row = header_row + i
        status = result.get("status", "SKIP")

        cells_data = [
            result.get("test_id", i),
            result.get("test_name", ""),
            result.get("screen", ""),
            status,
            f"{result.get('duration', 0):.2f}",
            result.get("error", "") or "—",
            os.path.basename(result.get("screenshot", "")) or "—",
            result.get("timestamp", ""),
        ]

        for col_idx, value in enumerate(cells_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = _alt_fill(i)
            cell.alignment = Alignment(horizontal="left" if col_idx > 2 else "center",
                                        vertical="center", wrap_text=True)
            cell.border = _thin_border()
            cell.font = Font(name="Calibri", color="DDDDDD", size=10)

            # Color the status cell
            if col_idx == 4:  # Status column
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
                if status == "PASS":
                    cell.fill = _pass_fill()
                elif status == "FAIL":
                    cell.fill = _fail_fill()
                else:
                    cell.fill = _skip_fill()

        ws.row_dimensions[row].height = 24

    _set_col_widths(ws, {
        "A": 6, "B": 40, "C": 30, "D": 12,
        "E": 14, "F": 45, "G": 40, "H": 22
    })

    # Freeze header
    ws.freeze_panes = ws[f"A{header_row + 1}"]


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 3 — SCREEN COVERAGE
# ─────────────────────────────────────────────────────────────────────────────

ALL_SCREENS = [
    ("Login Screen", "Authentication"),
    ("Forgot Password Screen", "Authentication"),
    ("Command Center", "Doctor Dashboard"),
    ("Patient Registration", "Patient Management"),
    ("Patient List / Directory", "Patient Management"),
    ("Dashboard (Central Console)", "Patient Monitoring"),
    ("Meal Planner (Intake Planner)", "Nutrition"),
    ("AI Food Scanner", "AI Features"),
    ("AI Face Analysis (Biometric Scan)", "AI Features"),
    ("AI Prediction Engine (Prognosis)", "AI Features"),
    ("Real-Time Analytics (Live Analytics)", "Analytics"),
    ("Weekly Progress Report", "Analytics"),
    ("Profile / User Settings & Logout", "Account"),
]


def _build_coverage_sheet(ws, results):
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "📱 Screen Coverage Map — Nutrixa App"
    title.font = _title_font(14)
    title.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers
    headers = ["#", "Screen Name", "Category", "Coverage", "Result"]
    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[header_row].height = 28

    # Build a quick lookup: screen_name → status
    status_map = {}
    for r in results:
        screen = r.get("screen", "")
        if screen not in status_map:
            status_map[screen] = r.get("status", "SKIP")

    for i, (screen_name, category) in enumerate(ALL_SCREENS, start=1):
        row = header_row + i
        # Try to find result for this screen
        matched_status = "SKIP"
        for r in results:
            if r.get("screen", "") == screen_name:
                matched_status = r.get("status", "SKIP")
                break

        coverage = "✅ Covered" if matched_status in ("PASS", "FAIL") else "⚠️ Not Run"

        row_data = [i, screen_name, category, coverage, matched_status]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = _alt_fill(i)
            cell.alignment = Alignment(horizontal="center" if col_idx in (1, 5) else "left",
                                        vertical="center")
            cell.border = _thin_border()
            cell.font = Font(name="Calibri", color="DDDDDD", size=10)

            if col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", bold=True,
                                  color="1DB954" if "Covered" in str(value) else "F57C00",
                                  size=10)
            if col_idx == 5:
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
                if matched_status == "PASS":
                    cell.fill = _pass_fill()
                elif matched_status == "FAIL":
                    cell.fill = _fail_fill()
                else:
                    cell.fill = _skip_fill()

        ws.row_dimensions[row].height = 24

    _set_col_widths(ws, {"A": 6, "B": 45, "C": 25, "D": 16, "E": 12})
